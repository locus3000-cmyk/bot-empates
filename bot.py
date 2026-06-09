import os, json, requests
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl import Workbook

# ===== CONFIG =====
API_KEY = os.environ["API_FOOTBALL_KEY"]
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
CHAT_ID = "691112681"
BASE = "https://v3.football.api-sports.io"
HEADERS = {"x-apisports-key": API_KEY}

PALABRAS_LIGAS = ["World Cup","Brazil","Argentina","Colombia","MLS","USA",
                  "Norway","Sweden","Finland","Japan","Korea"]
BANDA_MIN, BANDA_MAX = 2.5, 2.9
HORAS_ANTES = 3
MAX_ODDS_POR_PLAN = 40
AGENDA = "agenda.json"
EXCEL = "registro_empates.xlsx"

BOGOTA = timezone(timedelta(hours=-5))
ahora = datetime.now(BOGOTA)
hoy = ahora.strftime("%Y-%m-%d")
manana = (ahora + timedelta(days=1)).strftime("%Y-%m-%d")

def telegram(texto):
    requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                  data={"chat_id": CHAT_ID, "text": texto})

def cuota_empate(fid):
    ro = requests.get(f"{BASE}/odds", headers=HEADERS, params={"fixture": fid})
    cuotas = []
    for bloque in ro.json().get("response", []):
        for casa in bloque.get("bookmakers", []):
            for bet in casa.get("bets", []):
                if bet.get("name") == "Match Winner":
                    for v in bet.get("values", []):
                        if v.get("value") == "Draw":
                            try: cuotas.append(float(v["odd"]))
                            except: pass
    if not cuotas: return None
    return round(sum(cuotas)/len(cuotas), 2)

def nos_interesa(fx):
    t = ((fx["league"]["name"] or "")+" "+(fx["league"]["country"] or "")).lower()
    return any(p.lower() in t for p in PALABRAS_LIGAS)

def abrir_excel():
    if os.path.exists(EXCEL):
        wb = openpyxl.load_workbook(EXCEL); ws = wb.active
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Registro"
        ws.append(["Fecha","Partido","Liga","Hora (Col)","Cuota empate","Prob %",
                   "Resultado real","Marcador","Acerto?","ID"])
    return wb, ws

# ===== cargar agenda =====
if os.path.exists(AGENDA):
    with open(AGENDA, "r", encoding="utf-8") as f:
        agenda = json.load(f)
else:
    agenda = {"fecha_plan": "", "partidos": {}}

# ===== PARTE A: PLANEACION (una vez al dia) =====
if agenda.get("fecha_plan") != hoy:
    nuevos = 0
    for fecha in [hoy, manana]:
        r = requests.get(f"{BASE}/fixtures", headers=HEADERS,
                         params={"date": fecha, "timezone": "America/Bogota"})
        for fx in r.json().get("response", []):
            if not nos_interesa(fx): continue
            fid = str(fx["fixture"]["id"])
            if fid in agenda["partidos"]: continue
            if nuevos >= MAX_ODDS_POR_PLAN: continue
            cuota = cuota_empate(fid); nuevos += 1
            if cuota is None: continue
            agenda["partidos"][fid] = {
                "partido": f"{fx['teams']['home']['name']} vs {fx['teams']['away']['name']}",
                "liga": fx["league"]["name"],
                "kickoff": fx["fixture"]["date"],
                "cuota": cuota,
                "prob": round(100/cuota, 1),
                "en_banda": BANDA_MIN <= cuota <= BANDA_MAX,
                "alertado": False,
                "resuelto": False,
            }
    agenda["fecha_plan"] = hoy

    limite = ahora - timedelta(hours=12)
    for fid in list(agenda["partidos"].keys()):
        if datetime.fromisoformat(agenda["partidos"][fid]["kickoff"]) < limite:
            del agenda["partidos"][fid]

    cand = [p for p in agenda["partidos"].values() if p["en_banda"] and not p["alertado"]]
    telegram(f"📋 Agenda lista ({hoy}).\n"
             f"Partidos vigilados: {len(agenda['partidos'])}\n"
             f"Candidatos a empate (cuota {BANDA_MIN}-{BANDA_MAX}): {len(cand)}\n"
             f"Te aviso {HORAS_ANTES}h antes de cada uno.")

# ===== PARTE B: ENVIAR ALERTAS (cada corrida) =====
for fid, p in agenda["partidos"].items():
    if not p["en_banda"] or p["alertado"]: continue
    ko = datetime.fromisoformat(p["kickoff"])
    minutos = (ko - ahora).total_seconds() / 60
    if 0 < minutos <= HORAS_ANTES*60:
        telegram(f"⚽ ALERTA DE EMPATE\n\n"
                 f"{p['partido']}\n{p['liga']}\n"
                 f"{ko.strftime('%d/%m a las %H:%M')} (hora Colombia)\n"
                 f"Cuota empate: {p['cuota']}  ({p['prob']}%)\n\n"
                 f"Faltan ~{HORAS_ANTES} horas.")
        wb, ws = abrir_excel()
        ws.append([ko.strftime("%Y-%m-%d"), p["partido"], p["liga"],
                   ko.strftime("%H:%M"), p["cuota"], p["prob"], "", "", "", fid])
        wb.save(EXCEL)
        p["alertado"] = True

# ===== PARTE C: RESULTADOS (cada corrida) =====
for fid, p in agenda["partidos"].items():
    if not (p["alertado"] and p["en_banda"]) or p["resuelto"]: continue
    ko = datetime.fromisoformat(p["kickoff"])
    if (ahora - ko).total_seconds()/3600 < 2.5: continue
    rr = requests.get(f"{BASE}/fixtures", headers=HEADERS, params={"id": fid})
    resp = rr.json().get("response", [])
    if not resp: continue
    info = resp[0]
    if info["fixture"]["status"]["short"] not in ("FT","AET","PEN"): continue
    gl, gv = info["goals"]["home"], info["goals"]["away"]
    empato = (gl == gv); marcador = f"{gl}-{gv}"
    if os.path.exists(EXCEL):
        wb = openpyxl.load_workbook(EXCEL); ws = wb.active
        for fila in ws.iter_rows(min_row=2):
            if str(fila[9].value) == fid:
                fila[6].value = "EMPATE" if empato else "NO empate"
                fila[7].value = marcador
                fila[8].value = "SI" if empato else "NO"
        wb.save(EXCEL)
    p["resuelto"] = True
    telegram(f"📊 Resultado: {p['partido']}\n"
             f"Marcador {marcador} → {'ACERTAMOS ✅' if empato else 'No fue empate ❌'}")

# ===== guardar agenda =====
with open(AGENDA, "w", encoding="utf-8") as f:
    json.dump(agenda, f, ensure_ascii=False, indent=2)

print(f"Run OK {ahora.strftime('%Y-%m-%d %H:%M')}. Partidos en agenda: {len(agenda['partidos'])}")
